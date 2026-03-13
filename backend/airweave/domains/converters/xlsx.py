"""XLSX to markdown converter using openpyxl."""

import asyncio
from typing import Dict, List

from airweave.core.logging import logger
from airweave.domains.converters._base import BaseTextConverter
from airweave.domains.sync_pipeline.async_helpers import run_in_thread_pool
from airweave.domains.sync_pipeline.exceptions import EntityProcessingError, SyncFailureError


class XlsxConverter(BaseTextConverter):
    """Converts XLSX files to markdown using local openpyxl extraction."""

    async def convert_batch(self, file_paths: List[str]) -> Dict[str, str]:
        """Convert XLSX files to markdown text using openpyxl."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise SyncFailureError(
                "openpyxl package required for XLSX conversion but not installed"
            )

        logger.debug(f"Converting {len(file_paths)} XLSX files to markdown...")

        results = {}
        semaphore = asyncio.Semaphore(10)

        async def _convert_one(path: str):
            async with semaphore:
                try:
                    markdown = await self._extract_xlsx_to_markdown(path)

                    if not markdown or not markdown.strip():
                        logger.warning(f"XLSX extraction produced no content for {path}")
                        results[path] = None
                    else:
                        results[path] = markdown
                        logger.debug(f"Extracted XLSX: {path} ({len(markdown)} characters)")

                except EntityProcessingError as e:
                    logger.warning(f"XLSX conversion failed for {path}: {e}")
                    results[path] = None
                except Exception as e:
                    logger.error(f"Unexpected error converting XLSX {path}: {e}")
                    results[path] = None

        await asyncio.gather(*[_convert_one(p) for p in file_paths], return_exceptions=True)

        successful = sum(1 for r in results.values() if r)
        logger.debug(f"XLSX conversion complete: {successful}/{len(file_paths)} files successful")

        return results

    async def _extract_xlsx_to_markdown(self, xlsx_path: str) -> str:  # noqa: C901
        def _extract() -> str:  # noqa: C901
            from openpyxl import load_workbook

            try:
                wb = load_workbook(xlsx_path, data_only=False)
            except Exception as e:
                raise EntityProcessingError(f"Failed to open XLSX file {xlsx_path}: {e}")

            sheet_names = wb.sheetnames

            if not sheet_names:
                raise EntityProcessingError(f"XLSX file {xlsx_path} has no sheets")

            markdown_parts = []

            for sheet_name in sheet_names:
                sheet = wb[sheet_name]

                max_row = sheet.max_row
                max_col = sheet.max_column

                if max_row == 0 or max_col == 0:
                    logger.debug(f"Sheet '{sheet_name}' is empty, skipping")
                    continue

                markdown_parts.append(f"## Sheet: {sheet_name}\n")

                rows_data = []
                for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    row_values = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            row_values.append("")
                        else:
                            row_values.append(str(value))
                    rows_data.append(row_values)

                if not rows_data:
                    markdown_parts.append("*Empty sheet*\n")
                    continue

                if len(rows_data) > 1:
                    header = rows_data[0]
                    data_rows = rows_data[1:]

                    markdown_parts.append("| " + " | ".join(header) + " |")
                    markdown_parts.append("| " + " | ".join(["---"] * len(header)) + " |")

                    for row in data_rows:
                        padded_row = row + [""] * (len(header) - len(row))
                        markdown_parts.append("| " + " | ".join(padded_row[: len(header)]) + " |")
                else:
                    for value in rows_data[0]:
                        if value:
                            markdown_parts.append(f"- {value}")

                markdown_parts.append("")

            if not markdown_parts:
                raise EntityProcessingError(f"XLSX file {xlsx_path} has no extractable content")

            return "\n".join(markdown_parts)

        try:
            return await run_in_thread_pool(_extract)
        except EntityProcessingError:
            raise
        except Exception as e:
            raise EntityProcessingError(f"XLSX extraction failed for {xlsx_path}: {e}")
