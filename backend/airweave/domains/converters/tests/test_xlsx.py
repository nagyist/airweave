"""Tests for XlsxConverter — empty sheets error and valid data extraction."""

import os
import tempfile

import pytest

from airweave.domains.converters.xlsx import XlsxConverter


@pytest.fixture
def converter():
    return XlsxConverter()


@pytest.fixture
def temp_dir():
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


def _create_xlsx(path, sheets_data):
    """Create a test XLSX file.

    Args:
        path: File path to create
        sheets_data: dict of sheet_name -> list of rows (each row is a list of cell values)
    """
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for sheet_name, rows in sheets_data.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(path)


class TestXlsxConverter:

    @pytest.mark.asyncio
    async def test_valid_xlsx_returns_markdown(self, converter, temp_dir):
        file_path = os.path.join(temp_dir, "data.xlsx")
        _create_xlsx(
            file_path,
            {"Sheet1": [["Name", "Age"], ["Alice", 30], ["Bob", 25]]},
        )

        result = await converter.convert_batch([file_path])

        assert file_path in result
        assert result[file_path] is not None
        assert "Alice" in result[file_path]
        assert "Bob" in result[file_path]
        assert "| Name | Age |" in result[file_path]

    @pytest.mark.asyncio
    async def test_default_empty_workbook_still_extracts(self, converter, temp_dir):
        """XLSX with a default empty sheet still extracts a sheet header."""
        file_path = os.path.join(temp_dir, "empty.xlsx")
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        wb.save(file_path)

        result = await converter.convert_batch([file_path])

        assert file_path in result
        # openpyxl reports at least A1 even on an empty sheet
        assert result[file_path] is not None

    @pytest.mark.asyncio
    async def test_no_sheets_raises_error(self, converter, temp_dir):
        """XLSX with no sheet names → EntityProcessingError → None."""
        file_path = os.path.join(temp_dir, "no_sheets.xlsx")
        from openpyxl import Workbook

        wb = Workbook()
        wb.save(file_path)

        # Corrupt the file to have no sheets by removing them after save
        # We test via the error path more directly
        from unittest.mock import MagicMock, patch

        mock_wb = MagicMock()
        mock_wb.sheetnames = []

        with patch("openpyxl.load_workbook", return_value=mock_wb):
            result = await converter.convert_batch([file_path])

        assert file_path in result
        assert result[file_path] is None

    @pytest.mark.asyncio
    async def test_multi_sheet_xlsx(self, converter, temp_dir):
        file_path = os.path.join(temp_dir, "multi.xlsx")
        _create_xlsx(
            file_path,
            {
                "Users": [["Name"], ["Charlie"]],
                "Products": [["Item", "Price"], ["Widget", "9.99"]],
            },
        )

        result = await converter.convert_batch([file_path])

        assert file_path in result
        assert "Sheet: Users" in result[file_path]
        assert "Sheet: Products" in result[file_path]
        assert "Charlie" in result[file_path]
        assert "Widget" in result[file_path]

    @pytest.mark.asyncio
    async def test_nonexistent_file_returns_none(self, converter):
        result = await converter.convert_batch(["/nonexistent/file.xlsx"])
        assert result["/nonexistent/file.xlsx"] is None
